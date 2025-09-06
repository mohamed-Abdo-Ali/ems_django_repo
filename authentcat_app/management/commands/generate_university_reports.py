from django.core.management.base import BaseCommand
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from admin_app.models import Course, Major, Semester
from conttroll_app.models import (
    student_report_from_uivercity,
    Acdimaic_and_term_from_uivercity,
    student_courses_grads,
)
from taecher_app.models import Exam

# ====== الألوان والخطوط ======
RED_FILL = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")
HEADER_FILL = PatternFill(start_color="B4C6E7", end_color="D9E1F2", fill_type="solid")
ALT_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
WHITE_FILL = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

BOLD_FONT = Font(bold=True, size=12)
NORMAL_FONT = Font(bold=False, size=11)
HEADER_FONT = Font(bold=True, size=11)

CENTER = Alignment(horizontal="center", vertical="center")

# ====== حدود الخلايا ======
THIN_BORDER = Border(
    left=Side(border_style="thin", color="000000"),
    right=Side(border_style="thin", color="000000"),
    top=Side(border_style="thin", color="000000"),
    bottom=Side(border_style="thin", color="000000"),
)

PASS_MARK = 50  # درجة النجاح

def grade_text(average):
    if average >= 85:
        return "ممتاز"
    if average >= 75:
        return "جيد جدا"
    if average >= 65:
        return "جيد"
    if average >= 50:
        return "مقبول"
    return "راسب"

class Command(BaseCommand):
    help = "إنشاء تقارير الطلاب حسب الفصل والتخصص"

    def handle(self, *args, **options):
        acad = Acdimaic_and_term_from_uivercity.objects.last()
        acad_year = acad.Acdimaic_year if acad else ""
        acad_sem = acad.Acdimaic_year_semester if acad else ""

        semesters = student_report_from_uivercity.objects.values_list("semester_id", flat=True).distinct()
        majors = student_report_from_uivercity.objects.values_list("major", flat=True).distinct()

        wb = Workbook()
        default_sheet = wb.active
        default_sheet.title = "مؤقت"

        for sem_id in semesters:
            for major_name in majors:
                students = student_report_from_uivercity.objects.filter(
                    semester_id=sem_id, major=major_name
                ).order_by("name")

                if not students.exists():
                    continue

                try:
                    major_obj = Major.objects.get(name=major_name)
                except Major.DoesNotExist:
                    continue

                try:
                    semester = Semester.objects.get(id=sem_id)
                except Semester.DoesNotExist:
                    continue

                courses = Course.objects.filter(
                    semester=sem_id,
                    major=major_obj
                ).order_by("name")

                if not courses.exists():
                    continue

                # إنشاء ورقة جديدة
                sheet_name = f"{major_name[:20]}-الفصل {sem_id}"
                ws = wb.create_sheet(title=sheet_name)

                # تعيين RTL
                ws.sheet_view.rightToLeft = True

                num_cols = len(courses) + 6  # الأعمدة الكاملة بعد إضافة العمود الفارغ

                # ====== الرأس ======
                headers = [
                    f" التخصص: {major_name}",
                    f" السنة الدراسية: {acad_year}   |   الفصل: {acad_sem}",
                ]
                for idx, text in enumerate(headers, start=1):
                    ws.merge_cells(start_row=idx, start_column=1, end_row=idx, end_column=num_cols)
                    cell = ws.cell(row=idx, column=1, value=text)
                    cell.font = Font(size=14, bold=True, color="000000")
                    cell.alignment = CENTER
                    ws.row_dimensions[idx].height = 28

                ws.append([])

                # ====== رؤوس الأعمدة مع عمود فارغ قبل "م" ======
                header = [""] + ["م", "اسم الطالب"] + [c.name for c in courses] + ["المجموع", "المعدل", "التقدير"]
                ws.append(header)
                current_row = ws.max_row
                ws.row_dimensions[current_row].height = 27

                for col_idx in range(1, len(header) + 1):
                    if col_idx == 1:  # العمود الفارغ لا يطبق عليه أي تنسيقات
                        continue
                    cell = ws.cell(row=current_row, column=col_idx)
                    cell.font = BOLD_FONT
                    cell.alignment = CENTER
                    cell.fill = HEADER_FILL
                    cell.border = THIN_BORDER

                # ====== بيانات الطلاب ======
                counter = 1
                failed_count = 0    # عدد الطلاب الراسبين
                absent_count = 0    # عدد الطلاب الغائبين

                for student in students:
                    row = ["", counter, student.name]  # العمود الفارغ قبل العمود "م"
                    total = 0
                    failed_sub_count = 0
                    student_absent_flag = False  # علامة لمعرفة إذا الطالب غائب في أي مادة

                    for course in courses:
                        scg = student_courses_grads.objects.filter(student=student, course=course).first()
                        exams = Exam.objects.filter(course=course, exam_type=2).first()

                        if not exams:
                            row.append(0.0)
                        else:
                            if not scg or scg.total_mark is None:
                                row.append("غياب")
                                student_absent_flag = True
                            else:
                                val = float(scg.total_mark)
                                row.append(val)
                                total += val
                                if val < PASS_MARK:
                                    failed_sub_count += 1

                    # زيادة عدد الطلاب الغائبين مرة واحدة فقط
                    if student_absent_flag:
                        absent_count += 1

                    # زيادة عدد الطلاب الراسبين مرة واحدة فقط
                    if failed_sub_count > 0:
                        failed_count += 1

                    row.append(total)
                    avg = total / len(courses) if courses else 0
                    row.append(round(avg, 2))
                    row.append(f"مكمل/{failed_sub_count}" if failed_sub_count else grade_text(avg))

                    ws.append(row)
                    current_row = ws.max_row
                    ws.row_dimensions[current_row].height = 22

                    # تطبيق التنسيقات على كل الأعمدة ما عدا العمود الفارغ
                    for col_idx in range(1, len(row) + 1):
                        if col_idx == 1:  # العمود الفارغ
                            continue
                        cell = ws.cell(row=current_row, column=col_idx)
                        cell.font = NORMAL_FONT
                        cell.alignment = CENTER
                        cell.border = THIN_BORDER

                        # تلوين متناوب
                        if counter % 2 == 0:
                            cell.fill = ALT_FILL
                        else:
                            cell.fill = WHITE_FILL

                        # تلوين الغياب والرسوب بالأحمر
                        if col_idx >= 4 and (
                            cell.value == "غياب" or (isinstance(cell.value, (int, float)) and cell.value < PASS_MARK)
                        ):
                            cell.fill = RED_FILL

                    counter += 1

                # ====== الإحصائيات أفقي ======
                start_row = ws.max_row + 2
                stats = {
                    "إجمالي الطلاب": students.count(),
                    "عدد الراسبين": failed_count,
                    "عدد الغياب": absent_count,
                }

                col_offset = 1  # تعويض العمود الفارغ
                for col_idx, name in enumerate(stats.keys(), start=1 + col_offset):
                    cell = ws.cell(row=start_row, column=col_idx, value=name)
                    cell.font = HEADER_FONT
                    cell.fill = HEADER_FILL
                    cell.alignment = CENTER
                    cell.border = THIN_BORDER

                for col_idx, value in enumerate(stats.values(), start=1 + col_offset):
                    cell = ws.cell(row=start_row + 1, column=col_idx, value=value)
                    cell.font = HEADER_FONT
                    cell.fill = HEADER_FILL
                    cell.alignment = CENTER
                    cell.border = THIN_BORDER

                # ====== ضبط عرض الأعمدة ======
                for col_idx in range(1, ws.max_column + 1):
                    col_letter = get_column_letter(col_idx)
                    if col_letter == "A":
                        ws.column_dimensions[col_letter].width = 6.75
                    elif col_letter == "B":
                        ws.column_dimensions[col_letter].width = 6.75
                    elif col_letter == "C":
                        ws.column_dimensions[col_letter].width = 30
                    else:
                        max_length = 0
                        for row_cells in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
                            for cell in row_cells:
                                if cell.value:
                                    max_length = max(max_length, len(str(cell.value)))
                        ws.column_dimensions[col_letter].width = max_length + 5

        # حذف الورقة المؤقتة إذا تم إنشاء أوراق فعلية
        if len(wb.worksheets) > 1:
            wb.remove(default_sheet)
        else:
            default_sheet.title = "لا يوجد بيانات"

        out_path = "university_reports.xlsx"
        wb.save(out_path)
        self.stdout.write(self.style.SUCCESS(f"تم إنشاء الملف: {out_path}"))
